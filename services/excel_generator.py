from controllers.contract_controller import listar_contratos
from controllers.employee_controller import listar_empleados
from controllers.affiliation_controller import listar_afiliaciones
from typing import List, Dict, Any
from datetime import datetime

def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None

def _normalize_contract(c) -> Dict[str, Any]:
    if c is None:
        return {}
    if isinstance(c, dict):
        data = c.copy()
    else:
        try:
            data = c.__dict__.copy()
        except Exception:
            data = {}

    # normalizar claves comunes
    data.setdefault("id", data.get("id") or data.get("contract_id") or data.get("Id"))
    data.setdefault("inicio", data.get("inicio") or data.get("start_date"))
    data.setdefault("corte", data.get("corte") or data.get("end_date"))
    # añade aquí otros setdefault que necesites (empleado, tipo, valor_estimado, etc.)

    # detectar fecha efectiva para historial
    hist_keys = ["effective_date", "hist_effective_date", "fecha", "fecha_efectividad", "date", "inicio", "start_date"]
    for k in hist_keys:
        if k in data and data.get(k):
            p = _parse_date(data.get(k))
            if p:
                data["hist_effective_date"] = p
                break
    if "hist_effective_date" not in data:
        p = _parse_date(data.get("inicio") or data.get("start_date"))
        if p:
            data["hist_effective_date"] = p

    return data

def obtener_datos_contratos(include_all: bool = False) -> List[Dict[str, Any]]:
    """
    Devuelve lista de contratos normalizados.
    - Si include_all=True devuelve todas las filas normalizadas (sin agrupar).
    - Si include_all=False (por defecto) devuelve una fila por contrato (la más reciente).
    """
    raw = listar_contratos() or []
    norm = [_normalize_contract(r) for r in raw]

    if include_all:
        # limpiar campo auxiliar y devolver todas las filas
        for it in norm:
            it.pop("hist_effective_date", None)
        return norm

    # comportamiento por defecto: agrupar por id y devolver solo la fila más reciente
    grouped = {}
    for r in norm:
        cid = r.get("id")
        if cid is None:
            continue
        cur = grouped.get(cid)
        if not cur:
            grouped[cid] = r
            continue
        d_cur = cur.get("hist_effective_date")
        d_new = r.get("hist_effective_date")
        if d_new and d_cur:
            if d_new > d_cur:
                grouped[cid] = r
        elif d_new and not d_cur:
            grouped[cid] = r
        # si ninguno tiene fecha, mantiene el primero
    result = list(grouped.values())
    # limpiar campo auxiliar
    for it in result:
        it.pop("hist_effective_date", None)
    return result

def obtener_datos_empleados():
    empleados = listar_empleados() or []
    return [emp.__dict__ if not isinstance(emp, dict) else emp for emp in empleados]

def obtener_datos_afiliaciones():
    afiliaciones = listar_afiliaciones() or []
    return [a if isinstance(a, dict) else getattr(a, "__dict__", {}) for a in afiliaciones]

def export_sheets_to_excel(path: str, sheets: List[Dict[str, Any]], default_width: float = 15.0):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise ImportError("openpyxl no está instalado. Ejecuta: .venv\\Scripts\\python.exe -m pip install openpyxl") from e

    if not sheets:
        raise ValueError("No se proporcionaron hojas para exportar.")

    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.active)

    created_any = False
    for sheet in sheets:
        name = sheet.get("name", "Sheet1")
        rows = sheet.get("rows", []) or []
        columns: List[tuple] = sheet.get("columns", []) or []
        column_widths = sheet.get("column_widths", None)

        if not columns and not rows:
            continue

        ws = wb.create_sheet(title=name)
        created_any = True

        for col_idx, (key, header) in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            if isinstance(column_widths, (list, tuple)):
                w = column_widths[col_idx - 1] if len(column_widths) >= col_idx else default_width
            elif isinstance(column_widths, dict):
                w = column_widths.get(key, default_width)
            else:
                w = default_width
            ws.column_dimensions[get_column_letter(col_idx)].width = float(w)

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, (key, _) in enumerate(columns, start=1):
                val = ""
                if isinstance(row, dict):
                    val = row.get(key, "")
                else:
                    val = getattr(row, key, "")
                ws.cell(row=r_idx, column=c_idx, value=val)

        ws.sheet_state = 'visible'

    if not created_any:
        ws = wb.create_sheet(title="Empty")
        ws.cell(row=1, column=1, value="No hay datos para exportar")
        ws.sheet_state = 'visible'

    visible_index = None
    for i, sh in enumerate(wb.worksheets):
        state = getattr(sh, "sheet_state", "visible")
        if state == "visible":
            visible_index = i
            break

    if visible_index is None and wb.worksheets:
        wb.worksheets[0].sheet_state = 'visible'
        visible_index = 0

    try:
        if visible_index is not None:
            wb.active = int(visible_index)
    except Exception:
        try:
            wb._active_sheet_index = int(visible_index)
        except Exception:
            pass

    states = [(sh.title, getattr(sh, "sheet_state", "visible")) for sh in wb.worksheets]

    # log persistente para depuración (archivo temp en el proyecto)
    try:
        with open(r".\export_debug.log", "a", encoding="utf-8") as f:
            f.write(f"[export_debug] path={path}\n")
            f.write(f"[export_debug] sheets_provided={len(sheets)}, wb_sheets={len(wb.worksheets)}, created_any={created_any}, visible_index={visible_index}\n")
            f.write(f"[export_debug] states={states}\n\n")
    except Exception:
        pass

    try:
        wb.save(path)
    except Exception:
        import traceback
        with open(r".\export_debug.log", "a", encoding="utf-8") as f:
            f.write("[export_debug] Error al guardar workbook:\n")
            traceback.print_exc(file=f)
            f.write("\n\n")
        raise

